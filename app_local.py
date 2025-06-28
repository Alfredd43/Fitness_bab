import os
import json
import pandas as pd
import numpy as np
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, Response
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import plotly.express as px
import plotly.graph_objects as go
from plotly.utils import PlotlyJSONEncoder
from sklearn.ensemble import RandomForestClassifier
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, confusion_matrix, roc_auc_score
import pickle
import io
from datetime import datetime
import uuid

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-in-production'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('models', exist_ok=True)

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Simple user model for admin login
class User(UserMixin):
    def __init__(self, id):
        self.id = id

@login_manager.user_loader
def load_user(user_id):
    return User(user_id)

# Global variables for ML model and data
ml_model = None
scaler = None
current_data = None
clustering_results = None

def initialize_ml_model():
    """Initialize or load the ML model"""
    global ml_model, scaler
    
    # Try to load existing model from local storage
    try:
        with open('models/loan_recovery_model.pkl', 'rb') as f:
            ml_model = pickle.load(f)
        
        with open('models/scaler.pkl', 'rb') as f:
            scaler = pickle.load(f)
        print("Loaded existing model from local storage")
    except:
        # Create new model if none exists
        ml_model = RandomForestClassifier(n_estimators=100, random_state=42)
        scaler = StandardScaler()
        print("Created new ML model")

def save_model_locally():
    """Save ML model to local storage"""
    if ml_model and scaler:
        with open('models/loan_recovery_model.pkl', 'wb') as f:
            pickle.dump(ml_model, f)
        
        with open('models/scaler.pkl', 'wb') as f:
            pickle.dump(scaler, f)

def preprocess_data(df):
    """Preprocess the loan data"""
    # Handle missing values
    df = df.fillna(df.mean())
    
    # Convert categorical variables to numeric
    if 'region' in df.columns:
        df['region_encoded'] = pd.Categorical(df['region']).codes
    
    # Feature engineering
    df['overdue_ratio'] = df['overdue_days'] / df['loan_amount']
    df['credit_score_normalized'] = (df['credit_score'] - df['credit_score'].mean()) / df['credit_score'].std()
    
    # Select features for ML
    feature_columns = ['loan_amount', 'overdue_days', 'credit_score', 'overdue_ratio', 'credit_score_normalized']
    if 'region_encoded' in df.columns:
        feature_columns.append('region_encoded')
    
    return df, feature_columns

def train_model(df, feature_columns):
    """Train the ML model"""
    global ml_model, scaler
    
    # Prepare features and target (assuming 'recovered' column exists or create synthetic target)
    X = df[feature_columns]
    
    # Create synthetic target based on business rules
    # Higher credit score, lower overdue days, higher loan amount = more likely to recover
    df['recovery_score'] = (
        (df['credit_score'] / 850) * 0.4 +
        (1 - df['overdue_days'] / df['overdue_days'].max()) * 0.3 +
        (df['loan_amount'] / df['loan_amount'].max()) * 0.3
    )
    df['recovered'] = (df['recovery_score'] > 0.5).astype(int)
    
    y = df['recovered']
    
    # Split data
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    
    # Scale features
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)
    
    # Train model
    ml_model.fit(X_train_scaled, y_train)
    
    # Save model
    save_model_locally()
    
    return ml_model.score(X_test_scaled, y_test)

def predict_loans(df, feature_columns):
    """Make predictions on loan data"""
    global ml_model, scaler
    
    if ml_model is None or scaler is None:
        return None
    
    X = df[feature_columns]
    X_scaled = scaler.transform(X)
    
    predictions = ml_model.predict(X_scaled)
    probabilities = ml_model.predict_proba(X_scaled)
    
    return predictions, probabilities

def perform_clustering(df, feature_columns, n_clusters=3):
    """Perform K-means clustering"""
    X = df[feature_columns]
    X_scaled = scaler.transform(X)
    
    kmeans = KMeans(n_clusters=n_clusters, random_state=42)
    clusters = kmeans.fit_predict(X_scaled)
    
    return clusters, kmeans.cluster_centers_

def get_recovery_recommendations(prediction, probability, cluster):
    """Get recovery recommendations based on prediction and cluster"""
    recommendations = []
    
    if prediction == 1:  # Recoverable
        if probability >= 0.8:
            recommendations.extend([
                "Send gentle reminder email",
                "Offer 5% early payment discount",
                "Schedule follow-up call in 3 days"
            ])
        else:
            recommendations.extend([
                "Make immediate phone call",
                "Send formal payment reminder",
                "Offer payment plan options"
            ])
    else:  # Unrecoverable
        if probability >= 0.3:
            recommendations.extend([
                "Send urgent demand letter",
                "Consider debt restructuring",
                "Engage collection agency"
            ])
        else:
            recommendations.extend([
                "Prepare legal documentation",
                "Consider write-off",
                "Final demand notice"
            ])
    
    # Add cluster-specific recommendations
    if cluster == 0:
        recommendations.append("High-value customer - prioritize recovery")
    elif cluster == 1:
        recommendations.append("Medium-risk profile - standard procedures")
    else:
        recommendations.append("High-risk profile - aggressive recovery")
    
    return recommendations

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # Simple admin credentials (in production, use database)
        if username == 'admin' and password == 'admin123':
            user = User(username)
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid credentials')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if file and file.filename.endswith('.csv'):
        # Read CSV
        df = pd.read_csv(file)
        
        # Validate required columns
        required_columns = ['loan_amount', 'overdue_days', 'credit_score']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return jsonify({'error': f'Missing required columns: {missing_columns}'}), 400
        
        # Preprocess data
        df_processed, feature_columns = preprocess_data(df)
        
        # Train model
        accuracy = train_model(df_processed, feature_columns)
        
        # Make predictions
        predictions, probabilities = predict_loans(df_processed, feature_columns)
        
        # Perform clustering
        clusters, centers = perform_clustering(df_processed, feature_columns)
        
        # Add results to dataframe
        df_processed['prediction'] = predictions
        df_processed['probability'] = probabilities[:, 1]  # Probability of recovery
        df_processed['cluster'] = clusters
        
        # Create labels
        df_processed['recovery_label'] = df_processed['probability'].apply(
            lambda x: 'Recoverable' if x > 0.7 else 'Risky' if x > 0.4 else 'Unrecoverable'
        )
        
        # Store data globally
        global current_data, clustering_results
        current_data = df_processed
        clustering_results = {
            'clusters': clusters,
            'centers': centers,
            'feature_columns': feature_columns
        }
        
        # Save processed data locally
        df_processed.to_csv(f'uploads/predictions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv', index=False)
        
        # Save original file locally
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename)))
        
        return jsonify({
            'success': True,
            'accuracy': accuracy,
            'total_loans': len(df_processed),
            'recoverable': int((df_processed['prediction'] == 1).sum()),
            'unrecoverable': int((df_processed['prediction'] == 0).sum())
        })
    
    return jsonify({'error': 'Invalid file format'}), 400

@app.route('/predictions')
@login_required
def get_predictions():
    if current_data is None:
        return jsonify({'error': 'No data available'}), 400
    
    # Apply filters
    region_filter = request.args.get('region', '')
    min_overdue = request.args.get('min_overdue', 0, type=int)
    max_overdue = request.args.get('max_overdue', 999999, type=int)
    min_amount = request.args.get('min_amount', 0, type=float)
    max_amount = request.args.get('max_amount', float('inf'), type=float)
    min_credit = request.args.get('min_credit', 0, type=int)
    max_credit = request.args.get('max_credit', 999, type=int)
    
    filtered_data = current_data.copy()
    
    if region_filter and 'region' in filtered_data.columns:
        filtered_data = filtered_data[filtered_data['region'] == region_filter]
    
    filtered_data = filtered_data[
        (filtered_data['overdue_days'] >= min_overdue) &
        (filtered_data['overdue_days'] <= max_overdue) &
        (filtered_data['loan_amount'] >= min_amount) &
        (filtered_data['loan_amount'] <= max_amount) &
        (filtered_data['credit_score'] >= min_credit) &
        (filtered_data['credit_score'] <= max_credit)
    ]
    
    # Convert to JSON for frontend
    results = []
    for _, row in filtered_data.iterrows():
        recommendations = get_recovery_recommendations(
            row['prediction'], 
            row['probability'], 
            row['cluster']
        )
        
        results.append({
            'loan_amount': float(row['loan_amount']),
            'overdue_days': int(row['overdue_days']),
            'credit_score': int(row['credit_score']),
            'region': row.get('region', 'N/A'),
            'prediction': int(row['prediction']),
            'probability': float(row['probability']),
            'recovery_label': row['recovery_label'],
            'cluster': int(row['cluster']),
            'recommendations': recommendations
        })
    
    return jsonify(results)

@app.route('/analytics')
@login_required
def get_analytics():
    if current_data is None:
        return jsonify({'error': 'No data available'}), 400
    
    # Recovery rate by region
    if 'region' in current_data.columns:
        region_recovery = current_data.groupby('region')['prediction'].mean().reset_index()
        region_chart = px.bar(
            region_recovery, 
            x='region', 
            y='prediction',
            title='Recovery Rate by Region',
            color='prediction',
            color_continuous_scale='RdYlGn'
        )
    else:
        region_chart = None
    
    # Recovery rate by overdue days
    overdue_bins = pd.cut(current_data['overdue_days'], bins=5)
    overdue_recovery = current_data.groupby(overdue_bins)['prediction'].mean().reset_index()
    # Convert interval to string for JSON serialization
    overdue_recovery['overdue_days'] = overdue_recovery['overdue_days'].astype(str)
    overdue_chart = px.line(
        overdue_recovery,
        x='overdue_days',
        y='prediction',
        title='Recovery Rate by Overdue Days',
        markers=True
    )
    
    # Recovery rate by loan amount
    amount_bins = pd.cut(current_data['loan_amount'], bins=5)
    amount_recovery = current_data.groupby(amount_bins)['prediction'].mean().reset_index()
    # Convert interval to string for JSON serialization
    amount_recovery['loan_amount'] = amount_recovery['loan_amount'].astype(str)
    amount_chart = px.bar(
        amount_recovery,
        x='loan_amount',
        y='prediction',
        title='Recovery Rate by Loan Amount',
        color='prediction',
        color_continuous_scale='RdYlGn'
    )
    
    # Recovery label distribution
    label_dist = current_data['recovery_label'].value_counts()
    label_chart = px.pie(
        values=label_dist.values,
        names=label_dist.index,
        title='Distribution of Recovery Predictions',
        color_discrete_map={
            'Recoverable': '#28a745',
            'Risky': '#ffc107', 
            'Unrecoverable': '#dc3545'
        }
    )
    
    # Cluster distribution
    cluster_dist = current_data['cluster'].value_counts()
    cluster_chart = px.bar(
        x=cluster_dist.index,
        y=cluster_dist.values,
        title='Customer Clusters Distribution',
        color=cluster_dist.values,
        color_continuous_scale='Viridis'
    )
    
    # Credit score vs recovery probability
    credit_recovery = px.scatter(
        current_data,
        x='credit_score',
        y='probability',
        color='recovery_label',
        title='Credit Score vs Recovery Probability',
        color_discrete_map={
            'Recoverable': '#28a745',
            'Risky': '#ffc107',
            'Unrecoverable': '#dc3545'
        }
    )
    
    # Loan amount vs overdue days heatmap
    heatmap_data = current_data.groupby([
        pd.cut(current_data['loan_amount'], bins=5),
        pd.cut(current_data['overdue_days'], bins=5)
    ])['prediction'].mean().unstack()
    
    # Convert index and columns to strings for JSON serialization
    heatmap_data.index = heatmap_data.index.astype(str)
    heatmap_data.columns = heatmap_data.columns.astype(str)
    
    heatmap_chart = px.imshow(
        heatmap_data.values,
        x=heatmap_data.columns,
        y=heatmap_data.index,
        title='Recovery Rate Heatmap: Loan Amount vs Overdue Days',
        color_continuous_scale='RdYlGn'
    )
    
    # Convert charts to JSON
    charts = {
        'region': region_chart.to_json() if region_chart else None,
        'overdue': overdue_chart.to_json(),
        'amount': amount_chart.to_json(),
        'label': label_chart.to_json(),
        'cluster': cluster_chart.to_json(),
        'credit_recovery': credit_recovery.to_json(),
        'heatmap': heatmap_chart.to_json()
    }
    
    # Enhanced summary statistics
    summary = {
        'total_loans': len(current_data),
        'recoverable': int((current_data['prediction'] == 1).sum()),
        'unrecoverable': int((current_data['prediction'] == 0).sum()),
        'avg_probability': float(current_data['probability'].mean()),
        'avg_overdue_days': float(current_data['overdue_days'].mean()),
        'avg_loan_amount': float(current_data['loan_amount'].mean()),
        'avg_credit_score': float(current_data['credit_score'].mean()),
        'total_portfolio_value': float(current_data['loan_amount'].sum()),
        'recoverable_value': float(current_data[current_data['prediction'] == 1]['loan_amount'].sum()),
        'unrecoverable_value': float(current_data[current_data['prediction'] == 0]['loan_amount'].sum()),
        'recovery_rate': float((current_data['prediction'] == 1).mean()),
        'high_risk_loans': int((current_data['probability'] < 0.4).sum()),
        'medium_risk_loans': int(((current_data['probability'] >= 0.4) & (current_data['probability'] < 0.7)).sum()),
        'low_risk_loans': int((current_data['probability'] >= 0.7).sum())
    }
    
    return jsonify({'charts': charts, 'summary': summary})

@app.route('/export/<format>')
@login_required
def export_data(format):
    if current_data is None:
        return jsonify({'error': 'No data available'}), 400
    
    if format == 'csv':
        # Export to CSV
        csv_data = current_data.to_csv(index=False)
        return Response(
            csv_data,
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=loan_predictions.csv'}
        )
    
    elif format == 'excel':
        # Export to Excel
        try:
            from io import BytesIO
            import openpyxl
            from openpyxl import Workbook
            
            # Create Excel file
            output = BytesIO()
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Loan Predictions"
            
            # Add headers
            headers = ['Loan Amount', 'Overdue Days', 'Credit Score', 'Region', 
                      'Recovery Status', 'Probability', 'Cluster', 'Recommendations']
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # Add data
            for row, (_, item) in enumerate(current_data.iterrows(), 2):
                recommendations = get_recovery_recommendations(
                    item['prediction'], item['probability'], item['cluster']
                )
                worksheet.cell(row=row, column=1, value=item['loan_amount'])
                worksheet.cell(row=row, column=2, value=item['overdue_days'])
                worksheet.cell(row=row, column=3, value=item['credit_score'])
                worksheet.cell(row=row, column=4, value=item.get('region', 'N/A'))
                worksheet.cell(row=row, column=5, value=item['recovery_label'])
                worksheet.cell(row=row, column=6, value=item['probability'])
                worksheet.cell(row=row, column=7, value=item['cluster'])
                worksheet.cell(row=row, column=8, value='; '.join(recommendations))
            
            workbook.save(output)
            output.seek(0)
            
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': 'attachment; filename=loan_predictions.xlsx'}
            )
        except ImportError:
            return jsonify({'error': 'Excel export requires openpyxl package'}), 400
    
    elif format == 'json':
        # Export to JSON
        json_data = current_data.to_json(orient='records', indent=2)
        return Response(
            json_data,
            mimetype='application/json',
            headers={'Content-Disposition': 'attachment; filename=loan_predictions.json'}
        )
    
    else:
        return jsonify({'error': 'Unsupported format'}), 400

@app.route('/model_metrics')
@login_required
def get_model_metrics():
    if current_data is None:
        return jsonify({'error': 'No data available'}), 400
    
    # Calculate model performance metrics
    # Create synthetic target for evaluation
    y_true = (current_data['probability'] > 0.5).astype(int)
    y_pred = current_data['prediction']
    
    # Calculate metrics
    report = classification_report(y_true, y_pred, output_dict=True)
    conf_matrix = confusion_matrix(y_true, y_pred)
    auc_score = roc_auc_score(y_true, current_data['probability'])
    
    # Feature importance (if available)
    feature_importance = None
    if ml_model and hasattr(ml_model, 'feature_importances_'):
        feature_importance = dict(zip(
            ['loan_amount', 'overdue_days', 'credit_score', 'overdue_ratio', 'credit_score_normalized'],
            ml_model.feature_importances_
        ))
    
    metrics = {
        'accuracy': report['accuracy'],
        'precision': report['weighted avg']['precision'],
        'recall': report['weighted avg']['recall'],
        'f1_score': report['weighted avg']['f1-score'],
        'auc_score': auc_score,
        'confusion_matrix': conf_matrix.tolist(),
        'feature_importance': feature_importance,
        'total_samples': len(current_data),
        'positive_samples': int(y_true.sum()),
        'negative_samples': int((y_true == 0).sum())
    }
    
    return jsonify(metrics)

@app.route('/dashboard_data')
@login_required
def get_dashboard_data():
    if current_data is None:
        return jsonify({'error': 'No data available'}), 400
    
    # Basic charts (same as before)
    label_dist = current_data['recovery_label'].value_counts()
    recovery_chart = px.pie(
        values=label_dist.values,
        names=label_dist.index,
        title='Distribution of Recovery Predictions',
        color_discrete_map={
            'Recoverable': '#28a745',
            'Risky': '#ffc107', 
            'Unrecoverable': '#dc3545'
        }
    )
    
    cluster_dist = current_data['cluster'].value_counts()
    cluster_chart = px.bar(
        x=cluster_dist.index,
        y=cluster_dist.values,
        title='Customer Clusters Distribution',
        color=cluster_dist.values,
        color_continuous_scale='Viridis'
    )
    
    overdue_bins = pd.cut(current_data['overdue_days'], bins=5)
    overdue_recovery = current_data.groupby(overdue_bins)['prediction'].mean().reset_index()
    # Convert interval to string for JSON serialization
    overdue_recovery['overdue_days'] = overdue_recovery['overdue_days'].astype(str)
    overdue_chart = px.line(
        overdue_recovery,
        x='overdue_days',
        y='prediction',
        title='Recovery Rate by Overdue Days',
        markers=True
    )
    
    amount_bins = pd.cut(current_data['loan_amount'], bins=5)
    amount_recovery = current_data.groupby(amount_bins)['prediction'].mean().reset_index()
    # Convert interval to string for JSON serialization
    amount_recovery['loan_amount'] = amount_recovery['loan_amount'].astype(str)
    amount_chart = px.bar(
        amount_recovery,
        x='loan_amount',
        y='prediction',
        title='Recovery Rate by Loan Amount',
        color='prediction',
        color_continuous_scale='RdYlGn'
    )
    
    # Summary statistics
    summary = {
        'total_loans': len(current_data),
        'recoverable': int((current_data['prediction'] == 1).sum()),
        'unrecoverable': int((current_data['prediction'] == 0).sum()),
        'avg_probability': float(current_data['probability'].mean()),
        'avg_overdue_days': float(current_data['overdue_days'].mean()),
        'avg_loan_amount': float(current_data['loan_amount'].mean()),
        'avg_credit_score': float(current_data['credit_score'].mean()),
        'total_portfolio_value': float(current_data['loan_amount'].sum()),
        'recoverable_value': float(current_data[current_data['prediction'] == 1]['loan_amount'].sum()),
        'unrecoverable_value': float(current_data[current_data['prediction'] == 0]['loan_amount'].sum()),
        'recovery_rate': float((current_data['prediction'] == 1).mean()),
        'high_risk_loans': int((current_data['probability'] < 0.4).sum()),
        'medium_risk_loans': int(((current_data['probability'] >= 0.4) & (current_data['probability'] < 0.7)).sum()),
        'low_risk_loans': int((current_data['probability'] >= 0.7).sum())
    }
    
    # Convert data to list of dictionaries for JSON
    data_list = current_data.to_dict('records')
    
    return jsonify({
        'data': data_list,
        'summary': summary,
        'charts': {
            'recovery': recovery_chart.to_json(),
            'cluster': cluster_chart.to_json(),
            'overdue': overdue_chart.to_json(),
            'amount': amount_chart.to_json()
        }
    })

@app.route('/recommendations', methods=['GET', 'POST'])
@login_required
def recommendations():
    if request.method == 'GET':
        return render_template('recommendations.html')
    
    # Handle POST request for specific loan recommendations
    data = request.get_json()
    
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    # Extract loan data
    loan_amount = data.get('loan_amount', 0)
    overdue_days = data.get('overdue_days', 0)
    credit_score = data.get('credit_score', 0)
    recovery_label = data.get('recovery_label', 'Unknown')
    probability = data.get('probability', 0)
    cluster = data.get('cluster', 0)
    
    # Get recommendations based on the data
    recommendations = get_recovery_recommendations(
        prediction=1 if recovery_label == 'Recoverable' else 0,
        probability=probability,
        cluster=cluster
    )
    
    # Add specific recommendations based on loan characteristics
    if overdue_days > 90:
        recommendations.append("Consider legal action due to extended overdue period")
    
    if loan_amount > 100000:
        recommendations.append("High-value loan - prioritize recovery efforts")
    
    if credit_score < 600:
        recommendations.append("Low credit score - offer payment assistance programs")
    
    if probability < 0.3:
        recommendations.append("Very low recovery probability - consider write-off")
    
    return jsonify({
        'recommendations': recommendations,
        'loan_data': {
            'amount': loan_amount,
            'overdue_days': overdue_days,
            'credit_score': credit_score,
            'recovery_label': recovery_label,
            'probability': probability,
            'cluster': cluster
        }
    })

if __name__ == '__main__':
    initialize_ml_model()
    app.run(debug=True, host='0.0.0.0', port=5000) 